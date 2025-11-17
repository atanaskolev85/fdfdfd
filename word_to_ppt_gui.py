#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Word to PowerPoint Converter - GUI Application
Simple graphical interface for converting Word documents to PowerPoint
"""
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from word_to_ppt_converter import WordToPPTConverter


class WordToPPTApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Word to PowerPoint Converter")
        self.root.geometry("700x500")

        # Default template path
        self.template_path = "Project 742_051.pptx"
        self.word_path = ""
        self.output_dir = ""

        # Excel export options
        self.excel_export_enabled = tk.BooleanVar(value=False)
        self.excel_path = ""
        self.excel_sheet_name = tk.StringVar(value="Sheet1")

        # Auto-open folder option
        self.auto_open_folder = tk.BooleanVar(value=True)

        self.create_widgets()

    def create_widgets(self):
        # Title
        title_label = tk.Label(
            self.root,
            text="Word to PowerPoint Converter",
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=10)

        # Frame for file selections
        frame = tk.Frame(self.root)
        frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

        # Word file selection
        tk.Label(frame, text="Word Document:", font=("Arial", 10, "bold")).grid(
            row=0, column=0, sticky=tk.W, pady=5
        )
        self.word_label = tk.Label(frame, text="Не е избран файл", fg="gray")
        self.word_label.grid(row=0, column=1, sticky=tk.W, padx=10)
        tk.Button(frame, text="Избери Word файл", command=self.select_word_file).grid(
            row=0, column=2, padx=5
        )

        # Template selection
        tk.Label(frame, text="PowerPoint Бланка:", font=("Arial", 10, "bold")).grid(
            row=1, column=0, sticky=tk.W, pady=5
        )
        self.template_label = tk.Label(frame, text=self.template_path, fg="blue")
        self.template_label.grid(row=1, column=1, sticky=tk.W, padx=10)
        tk.Button(frame, text="Промени бланка", command=self.select_template).grid(
            row=1, column=2, padx=5
        )

        # Output directory
        tk.Label(frame, text="Запази в:", font=("Arial", 10, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=5
        )
        self.output_label = tk.Label(frame, text="Същата папка като Word файла", fg="gray")
        self.output_label.grid(row=2, column=1, sticky=tk.W, padx=10)
        tk.Button(frame, text="Избери папка", command=self.select_output_dir).grid(
            row=2, column=2, padx=5
        )

        # Separator
        ttk.Separator(frame, orient='horizontal').grid(row=3, column=0, columnspan=3, sticky='ew', pady=10)

        # Excel export checkbox
        self.excel_checkbox = tk.Checkbutton(
            frame,
            text="Експортирай данни в Excel файл",
            variable=self.excel_export_enabled,
            font=("Arial", 10, "bold"),
            command=self.toggle_excel_options
        )
        self.excel_checkbox.grid(row=4, column=0, columnspan=3, sticky=tk.W, pady=5)

        # Excel file selection
        tk.Label(frame, text="Excel файл:", font=("Arial", 10)).grid(
            row=5, column=0, sticky=tk.W, pady=5
        )
        self.excel_label = tk.Label(frame, text="Не е избран файл", fg="gray")
        self.excel_label.grid(row=5, column=1, sticky=tk.W, padx=10)
        self.excel_btn = tk.Button(frame, text="Избери Excel файл", command=self.select_excel_file, state=tk.DISABLED)
        self.excel_btn.grid(row=5, column=2, padx=5)

        # Sheet name selection (dropdown)
        tk.Label(frame, text="Избери sheet:", font=("Arial", 10)).grid(
            row=6, column=0, sticky=tk.W, pady=5
        )
        self.sheet_combo = ttk.Combobox(frame, textvariable=self.excel_sheet_name, width=28, state=tk.DISABLED)
        self.sheet_combo.grid(row=6, column=1, sticky=tk.W, padx=10)

        # Separator
        ttk.Separator(frame, orient='horizontal').grid(row=7, column=0, columnspan=3, sticky='ew', pady=10)

        # Auto-open folder checkbox
        self.auto_open_checkbox = tk.Checkbutton(
            frame,
            text="Отваряй папката автоматично след конвертиране",
            variable=self.auto_open_folder,
            font=("Arial", 9)
        )
        self.auto_open_checkbox.grid(row=8, column=0, columnspan=3, sticky=tk.W, pady=5)

        # Progress text
        tk.Label(self.root, text="Статус:", font=("Arial", 10, "bold")).pack(pady=10)
        self.progress_text = tk.Text(self.root, height=6, width=70)
        self.progress_text.pack(pady=5)

        # Convert button
        self.convert_btn = tk.Button(
            self.root,
            text="Конвертирай",
            command=self.convert,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 12, "bold"),
            width=20,
            height=2
        )
        self.convert_btn.pack(pady=20)

    def log(self, message):
        """Add message to progress text"""
        self.progress_text.insert(tk.END, message + "\n")
        self.progress_text.see(tk.END)
        self.root.update()

    def select_word_file(self):
        """Select Word document"""
        filepath = filedialog.askopenfilename(
            title="Избери Word документ",
            filetypes=[("Word Documents", "*.docx"), ("All files", "*.*")]
        )
        if filepath:
            self.word_path = filepath
            self.word_label.config(text=os.path.basename(filepath), fg="blue")
            self.log(f"Избран Word файл: {os.path.basename(filepath)}")

    def select_template(self):
        """Select PowerPoint template"""
        filepath = filedialog.askopenfilename(
            title="Избери PowerPoint бланка",
            filetypes=[("PowerPoint", "*.pptx"), ("All files", "*.*")]
        )
        if filepath:
            self.template_path = filepath
            self.template_label.config(text=os.path.basename(filepath))
            self.log(f"Избрана бланка: {os.path.basename(filepath)}")

    def select_output_dir(self):
        """Select output directory"""
        dirpath = filedialog.askdirectory(title="Избери папка за запис")
        if dirpath:
            self.output_dir = dirpath
            self.output_label.config(text=dirpath, fg="blue")
            self.log(f"Папка за запис: {dirpath}")

    def toggle_excel_options(self):
        """Enable/disable Excel export options"""
        if self.excel_export_enabled.get():
            self.excel_btn.config(state=tk.NORMAL)
            self.sheet_combo.config(state="readonly")
        else:
            self.excel_btn.config(state=tk.DISABLED)
            self.sheet_combo.config(state=tk.DISABLED)

    def select_excel_file(self):
        """Select Excel file"""
        filepath = filedialog.askopenfilename(
            title="Избери Excel файл",
            filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")]
        )
        if filepath:
            self.excel_path = filepath
            self.excel_label.config(text=os.path.basename(filepath), fg="blue")
            self.log(f"Избран Excel файл: {os.path.basename(filepath)}")

            # Load sheet names from Excel file
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                # Update combobox with sheet names
                self.sheet_combo['values'] = sheet_names

                # Select first sheet by default
                if sheet_names:
                    self.excel_sheet_name.set(sheet_names[0])
                    self.log(f"  Намерени {len(sheet_names)} sheets: {', '.join(sheet_names)}")

            except Exception as e:
                self.log(f"  Грешка при зареждане на sheets: {e}")
                messagebox.showwarning("Внимание", f"Не успях да заредя sheet имената:\n{e}")

    def convert(self):
        """Perform conversion"""
        # Validate inputs
        if not self.word_path:
            messagebox.showerror("Грешка", "Моля избери Word документ!")
            return

        if not os.path.exists(self.template_path):
            messagebox.showerror("Грешка", f"PowerPoint бланката не съществува:\n{self.template_path}")
            return

        # Validate Excel export if enabled
        if self.excel_export_enabled.get():
            if not self.excel_path:
                messagebox.showerror("Грешка", "Моля избери Excel файл за експорт!")
                return
            if not os.path.exists(self.excel_path):
                messagebox.showerror("Грешка", f"Excel файлът не съществува:\n{self.excel_path}")
                return
            if not self.excel_sheet_name.get():
                messagebox.showerror("Грешка", "Моля въведи име на sheet!")
                return

        # Clear progress
        self.progress_text.delete(1.0, tk.END)
        self.log("Започване на конвертиране...")

        try:
            # Create converter
            converter = WordToPPTConverter(self.word_path, self.template_path)

            # Determine output directory
            output_dir = self.output_dir if self.output_dir else None

            # Set Excel export options if enabled
            if self.excel_export_enabled.get():
                converter.set_excel_export(
                    self.excel_path,
                    self.excel_sheet_name.get()
                )
                self.log(f"Excel експорт активиран: {os.path.basename(self.excel_path)}")

            # Convert (this will extract data and update PowerPoint)
            output_path = converter.convert(output_dir)

            # Show extracted data after conversion
            self.log("\nИзвлечени данни:")
            for key, value in converter.data.items():
                self.log(f"  {key}: {value}")

            # Success
            self.log("\n✓ Готово!")
            self.log(f"✓ Файлът е записан: {output_path}")

            # Show success message
            messagebox.showinfo(
                "Успех!",
                f"Конвертирането е успешно!\n\nФайлът е записан в:\n{output_path}",
            )

            # Open output directory if option is enabled
            if self.auto_open_folder.get():
                try:
                    os.startfile(os.path.dirname(output_path))
                except:
                    pass

        except Exception as e:
            self.log(f"\n✗ Грешка: {str(e)}")
            messagebox.showerror("Грешка", f"Грешка при конвертиране:\n\n{str(e)}")


def main():
    root = tk.Tk()
    app = WordToPPTApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
